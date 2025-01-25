import os
import nltk
import nest_asyncio
nest_asyncio.apply()
import pickle
import argparse
from llama_index.core import SimpleDirectoryReader
from llama_index.core.readers.base import BaseReader
from llama_index.readers.file import UnstructuredReader
from llama_parse import LlamaParse
from config import config
from markdownify import markdownify as md
from tqdm import tqdm
from pathlib import Path
from typing import Any, Dict, List, Optional
from llama_index.core.schema import Document
import pandas as pd
from openpyxl import load_workbook
import hashlib

class XlsxReader(BaseReader):
    def __init__(
        self,
    ) -> None:
        super().__init__()

    def xlsx_load(self,file: Path) -> str:
        wb = load_workbook(file)
        # 获取所有工作表的名称
        sheet_names = wb.sheetnames

        markdown_menu = ""

        # 遍历每一个工作表
        for sheet_name in sheet_names:
            sub_wb = wb[sheet_name]
            merged_cells = list(sub_wb.merged_cells.ranges)  # 转换为列表
            
            # 遍历每一个合并单元格
            for merged_cell in merged_cells:
                min_row, max_row = merged_cell.min_row, merged_cell.max_row
                min_col, max_col = merged_cell.min_col, merged_cell.max_col
                
                # 获取合并单元格的值
                cell_value = sub_wb.cell(row=min_row, column=min_col).value
                
                # 解除合并单元格
                sub_wb.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
                
                # 将值填充到之前合并单元格的所有单元格中
                for col in range(min_col, max_col + 1):
                    for row in range(min_row, max_row + 1):
                        sub_wb.cell(row=row, column=col, value=cell_value)

            data = wb[sheet_name].values
            columns = next(data)[0:]  # 获取第一行作为列名
            df = pd.DataFrame(data, columns=columns)
            
            # 处理 DataFrame 中的回车符
            df = df.applymap(lambda x: str(x).replace('\n', ' ') if isinstance(x, str) else x)
            
            # 去掉全为空值的行和列
            df.dropna(how='all', inplace=True)
            df.dropna(axis=1, how='all', inplace=True)

            # 去掉全为 None 的行和列
            df = df.applymap(lambda x: None if x == 'None' else x)
            df.dropna(how='all', inplace=True)
            df.dropna(axis=1, how='all', inplace=True)

            # 将 DataFrame 转换为 Markdown 格式
            markdown_output = df.to_markdown(index=False)
            markdown_menu += sheet_name + "菜单：\n"
            markdown_menu += markdown_output
            markdown_menu += "\n\n\n\n"

        return markdown_menu


    def load_data(
        self, file: Path, extra_info: Optional[Dict] = None
    ) -> List[Document]:
        docs = []
        metadata = {
            "file_name": file.name,
            "file_path": str(file),
        }
        if extra_info is not None:
            metadata.update(extra_info)

        return [Document(text=self.xlsx_load(file), metadata=metadata or {})]

def hash_file(filename):
    h = hashlib.sha256()
    with open(filename, "rb") as file:
        while True:
            chunk = file.read(h.block_size)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def hash_directory(directory):
    all_hashes = ""
    for root, _, files in os.walk(directory):
        for filename in files:
            filepath = os.path.join(root, filename)
            file_hash = hash_file(filepath)
            all_hashes += file_hash
    final_hash = hashlib.sha256(all_hashes.encode("utf-8")).hexdigest()
    return final_hash


def update_data(data_dir):
    reader = UnstructuredReader()

    documents_path = "../datapool/RAG_data/pdf.pkl"

    xlsx_reader = XlsxReader()
    llama_parse_api_key = "llx-ruUEWvib0ZlDnk75bwLWfvNh1x117Kl2Z6ecpPL0tLLnJMdK"
    pdf_parser = LlamaParse(
        api_key=llama_parse_api_key,
        result_type="markdown",
        verbose=True,
    )
    documents = SimpleDirectoryReader(
        data_dir,
        recursive=True,
        required_exts=[ ".pdf", ".csv", ".xlsx"],
        file_extractor={
            ".pdf": pdf_parser,
            ".csv": reader,
            ".xlsx": xlsx_reader,
        },
    ).load_data()

    for doc in documents:
        if doc.metadata["file_type"] == "text/html":
            with open(doc.metadata["file_path"], "r") as f:
                html = f.read()
            try:
                doc.text = md(html)
            except:
                print(f"fail trans to md:{doc.metadata['file_path']}")

    with open(documents_path, "wb") as f:
        pickle.dump(documents, f)
    print(f"Documents stored in {documents_path}")
    print("Length of documents:", len(documents))
    return documents


def main(data_dir=None):

    update_data(data_dir)

if __name__ == "__main__":
    main(config.data_dir)
